import openpyxl as xl
from openpyxl.utils import column_index_from_string as l2i

wb = xl.load_workbook("FY2023 CDS.xlsx")
sh = wb["results"]

# amount is column H
# state is column I

totals = {}

for row in range(8, sh.max_row):
    amount = sh.cell(column=l2i("H"), row=row).value
    states = sh.cell(column=l2i("I"), row=row).value

    if states == None:
        continue

    xs = states.split(',')
    for state in xs:
        state = state.strip()
        amt = totals.get(state, 0)
        amt += int(amount) / len(xs)
        totals[state] = amt

out_wb = xl.Workbook()
sh = out_wb.active
sh.title = "CDS Totals by State"

sh["A1"] = "State"
sh["B1"] = "Total"

ii = 2
for st in sorted(totals.keys()):
    print(st, totals[st])
    sh[f"A{ii}"] = st
    sh[f"B{ii}"] = totals[st]
    ii += 1

out_wb.save("output.xlsx")







